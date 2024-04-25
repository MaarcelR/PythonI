# Instando a biblioteca openpyxl
# pip install openpyxl
# import openpyxl
# from openpyxl import load_workbook
# arquivo='planilha.xlsx'
# workbook=load_workbook(arquivo)
# planilha=workbook.active

# for linha in planilha.iter_rows():
#     for celula in linha:
#         print(celula.value, end='\t')
#     print(linha)
#  outra forma de fazer 
# for linha in planilha.iter_rows():
#     for celula in linha:
#         print(str(celula.value).ljust(18), end='\t') ## Ljust justifica as colunas deixa mais apresentavel
#     print(celula)


### EXIBIR UMA INFORACAO USANDO REFERENCIA DE CELULA

import openpyxl
planilha=openpyxl.load_workbook("planilha.xlsx")
aba=planilha["PRODUTOS"]
print(aba['a2'].value)
print(aba['b2'].value)






